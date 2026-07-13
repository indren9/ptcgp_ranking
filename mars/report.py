# mars/report.py
from __future__ import annotations

from collections import OrderedDict
from typing import Dict, Iterable, Optional, Tuple
from pathlib import Path
import logging

import numpy as np
import pandas as pd

from reporting.excel import write_excel_versioned_styled  # writer con styling/CF

LOGGER = logging.getLogger("ptcgp")
_PCT = 100.0


# ──────────────────────────────────────────────────────────────────────────────
# Helpers base
# ──────────────────────────────────────────────────────────────────────────────
def _sanitize_sheet_name(name: str) -> str:
    """Excel: max 31 char, no []:*?/\\ ."""
    bad = set('[]:*?/\\')
    safe = "".join(ch for ch in str(name) if ch not in bad).strip()
    return (safe or "Sheet")[:31]


def _ensure_axis(df: pd.DataFrame, axis: Iterable[str]) -> pd.DataFrame:
    axis = list(axis)
    return df.reindex(index=axis, columns=axis)


def _wr_real_from_score(score_flat: pd.DataFrame, axis: Iterable[str]) -> pd.DataFrame:
    """Matrice WR_real_% T×T (direzionale A→B) da score_latest (flat)."""
    need = {"Deck A", "Deck B", "WR_dir"}
    missing = need.difference(score_flat.columns)
    if missing:
        raise ValueError(f"score_flat is missing columns: {sorted(missing)}")

    wr_pivot = (
        score_flat.astype({"Deck A": "string", "Deck B": "string"})
        .pivot(index="Deck A", columns="Deck B", values="WR_dir")
    )
    wr_pivot = _ensure_axis(wr_pivot, axis)
    np.fill_diagonal(wr_pivot.values, np.nan)
    return wr_pivot


def _counts_from_score(score_flat: pd.DataFrame, axis: Iterable[str]) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Matrici W e L T×T da score_latest (flat)."""
    need = {"Deck A", "Deck B", "W", "L"}
    missing = need.difference(score_flat.columns)
    if missing:
        raise ValueError(f"score_flat is missing columns: {sorted(missing)}")

    W = (
        score_flat.pivot(index="Deck A", columns="Deck B", values="W")
        .reindex(index=axis, columns=axis)
        .fillna(0.0).astype(float)
    )
    L = (
        score_flat.pivot(index="Deck A", columns="Deck B", values="L")
        .reindex(index=axis, columns=axis)
        .fillna(0.0).astype(float)
    )
    np.fill_diagonal(W.values, np.nan)
    np.fill_diagonal(L.values, np.nan)
    return W, L


def _posterior_from_wr_n(
    wr_dir_pct: pd.DataFrame,
    n_dir: pd.DataFrame,
    *,
    mu: float,
    K: float,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Posteriori Beta su ogni cella A→B usando p_obs e N:
      a = mu*K + W,  b = (1-mu)*K + L, con W = p_obs*N, L = (1-p_obs)*N
    Restituisce: p_hat_% (T×T) e SE_dir_% (T×T) = 100*sqrt(Var[Beta(a,b)]).
    """
    if set(wr_dir_pct.columns) != set(n_dir.columns) or set(wr_dir_pct.index) != set(n_dir.index):
        raise ValueError("wr_dir_pct and n_dir do not share the same axis.")

    p_obs = wr_dir_pct / _PCT
    N = n_dir.astype(float)

    a = mu * K + (p_obs * N)
    b = (1.0 - mu) * K + ((1.0 - p_obs) * N)
    denom = a + b

    p_hat = a / denom
    var = (a * b) / (denom.pow(2) * (denom + 1.0))
    se = np.sqrt(var)

    return (p_hat * _PCT), (se * _PCT)


def _se_binom_from_wr_n(wr_dir_pct: pd.DataFrame, n_dir: pd.DataFrame) -> pd.DataFrame:
    """SE binomiale della proporzione osservata: 100·sqrt(p(1-p)/N_dir). NaN se N_dir==0."""
    p = wr_dir_pct / _PCT
    N = n_dir.astype(float)
    with np.errstate(divide="ignore", invalid="ignore"):
        se = np.sqrt(np.clip(p * (1.0 - p) / N, 0.0, None))
    return se * _PCT


def _weights_row_for_A(p_blend: pd.Series, A: str, axis: Iterable[str]) -> pd.Series:
    """Rinormalizza i pesi meta-blend su B≠A; somma=1."""
    axis = list(axis)
    w = p_blend.reindex(axis).astype(float).copy()
    w.loc[A] = 0.0
    s = w.sum()
    if s <= 0:
        n_opp = max(len(axis) - 1, 1)
        w = pd.Series({b: (0.0 if b == A else 1.0 / n_opp) for b in axis}, dtype=float)
    else:
        w /= s
    return w


# ──────────────────────────────────────────────────────────────────────────────
# Per-deck tables + legend (data-only; no I/O)
# ──────────────────────────────────────────────────────────────────────────────
def make_pairs_by_deck_tables(
    *,
    filtered_wr: pd.DataFrame,         # T×T, % direzionali A→B, diag NaN
    n_dir: pd.DataFrame,               # T×T, W+L direzionali, diag NaN
    p_blend: pd.Series,                # pesi meta-blend su asse (stessi della MAS)
    K_used: float,
    score_flat: Optional[pd.DataFrame] = None,  # se presente, WR_real_% e conteggi da qui
    mu: float = 0.5,
    include_posterior_se: bool = False,        # default OFF come richiesto
    include_binom_se: bool = True,
    gamma: Optional[float] = None,
    include_counts: bool = True,               # add W, L, N columns when available
    sort_by: Optional[str] = None,             # ignored when global_order is passed
    global_order: Optional[Iterable[str]] = None,  # ordine righe fisso (es. ranking)
    include_self_row: bool = True,             # inserisce riga del deck A come "Mirror"
    mirror_label: str = "Mirror",
    include_weight_col: bool = False,          # w_A(B)_% (default OFF)
    include_mas_contrib_col: bool = False,     # MAS_contrib_pp (default OFF)
) -> Tuple["OrderedDict[str, pd.DataFrame]", pd.DataFrame, Dict]:
    """
    Costruisce:
      - sheets_by_deck: {A -> DataFrame}
      - legend_df: complete legend used by the PNG banner
      - meta: info varie
    """
    axis: list[str] = list(filtered_wr.columns)
    filtered_wr = filtered_wr.reindex(index=axis, columns=axis)
    n_dir = n_dir.reindex(index=axis, columns=axis)
    p_blend = p_blend.reindex(axis).astype(float)

    # WR reale e (opz) conteggi W/L
    if score_flat is not None:
        wr_real_pct = _wr_real_from_score(score_flat, axis)
        W_mat, L_mat = _counts_from_score(score_flat, axis) if include_counts else (None, None)
    else:
        wr_real_pct = filtered_wr.copy()
        W_mat, L_mat = (None, None)

    # Posteriori + SE posterior (Beta)
    p_hat_pct, se_post_pct = _posterior_from_wr_n(wr_real_pct, n_dir, mu=mu, K=K_used)

    # SE binomiale
    se_binom_pct = _se_binom_from_wr_n(wr_real_pct, n_dir)

    # Ordine fisso righe (uguale per tutti i fogli)
    order = list(global_order) if global_order is not None else list(axis)

    sheets: "OrderedDict[str, pd.DataFrame]" = OrderedDict()
    for A in axis:
        rows = []
        need_weights = include_weight_col or include_mas_contrib_col
        w_row = _weights_row_for_A(p_blend, A, axis) if need_weights else None

        for B in order:
            # Riga Mirror
            if B == A and include_self_row:
                row = {"Opponent": mirror_label}
                if include_counts and (W_mat is not None) and (L_mat is not None):
                    row.update({"W": pd.NA, "L": pd.NA, "N": pd.NA})
                row.update({"WR_real_%": np.nan, "p_hat_%": np.nan})
                if include_posterior_se: row["SE_dir_%"] = np.nan
                if include_binom_se:     row["SE_binom_%"] = np.nan
                row["gap_pp"] = np.nan
                if include_weight_col:      row["w_A(B)_%"] = 0.0
                if include_mas_contrib_col: row["MAS_contrib_pp"] = 0.0
                rows.append(row)
                continue
            if B == A and not include_self_row:
                continue

            # Riga normale A→B
            row = {"Opponent": B}
            if include_counts and (W_mat is not None) and (L_mat is not None):
                Wv = W_mat.loc[A, B]; Lv = L_mat.loc[A, B]
                Nv = (Wv + Lv) if pd.notna(Wv) and pd.notna(Lv) else np.nan
                row.update({"W": Wv, "L": Lv, "N": Nv})

            wr_ab = wr_real_pct.loc[A, B]
            ph_ab = p_hat_pct.loc[A, B]
            row.update({"WR_real_%": wr_ab, "p_hat_%": ph_ab})
            if include_posterior_se: row["SE_dir_%"] = se_post_pct.loc[A, B]
            if include_binom_se:     row["SE_binom_%"] = se_binom_pct.loc[A, B]

            row["gap_pp"] = (ph_ab - wr_ab)

            if need_weights:
                wB = float(w_row.loc[B]) * _PCT
                if include_weight_col:
                    row["w_A(B)_%"] = wB
                if include_mas_contrib_col:
                    row["MAS_contrib_pp"] = (wB * ph_ab / _PCT)

            rows.append(row)

        df = pd.DataFrame(rows)

        # If global_order is not used, allow per-sheet sort_by.
        if global_order is None and sort_by and sort_by in df.columns:
            df = df.sort_values(by=sort_by, ascending=False, kind="mergesort").reset_index(drop=True)

        # Types/rounding.
        for col in df.columns:
            if col in ("W", "L", "N"):
                df[col] = pd.to_numeric(df[col], errors="coerce").astype("Int64")
            elif col != "Opponent":
                df[col] = pd.to_numeric(df[col], errors="coerce").round(2)

        sheets[A] = df

    # ---------- LEGEND ----------
    # Cover: What this is.
    catchy = (f"This workbook presents the Top {len(axis)} deck ranking in '01_Summary'. "
              "Then it provides one sheet per deck A: A against every other deck, in ranking order.")
    # Per-deck sheet columns, limited to the columns that are actually present.
    per_deck_rows = [
        ("Units", "Percentages use 2 decimals. 'pp' means percentage points. W/L/N are directional counts (A->B)."),
        ("W,L,N", "W=wins, L=losses, N=W+L (A->B; ties are excluded from observed WR)."),
        ("WR_real_%", "Observed win rate: 100*W/(W+L), after aliases and filters."),
        ("p_hat_%", "Low-data adjusted win rate: 100*(W+mu*K)/(W+L+K), with mu=0.5 and data-selected K."),
        ("SE_binom_%", "Standard error of observed WR: 100*sqrt(p*(1-p)/N_dir). NaN if N_dir=0."),
        ("gap_pp", "Applied correction: p_hat_% - WR_real_% (positive = raises; negative = lowers)."),
        ("Mirror", "Deck A against itself, included for readability with empty metric fields."),
    ]
    if include_posterior_se:
        per_deck_rows.insert(4, ("SE_dir_%", "Uncertainty of the adjusted estimate (standard deviation), defined even with small N."))
    if include_weight_col:
        per_deck_rows.insert(-1, ("w_A(B)_%", "Weight of deck B in A's average (meta-blend weights; sum to about 100 over B!=A)."))
    if include_mas_contrib_col:
        per_deck_rows.insert(-1, ("MAS_contrib_pp", "Contribution of B to A's expected performance: w_A(B)% * p_hat_% / 100. The sum reconstructs A's MAS_%."))

    legend_df = pd.DataFrame(
        [("What this is", catchy)] + per_deck_rows
        + [("Run parameters", f"T={len(axis)}; mu={mu}; K_used={K_used}" + (f"; gamma={gamma}" if gamma is not None else "")),
           ("Conventions", "All per-deck sheets use the same row order as the ranking.")],
        columns=["Field", "Description"]
    )
    legend_df["Color"] = pd.NA

    # Ranking section (01_Summary).
    ranking_legend_rows = [
        ("Deck", "Deck name after alias unification."),
        ("Score_%", "Final score (0-100): blend of LB_% (conservative estimate) and BT_% (head-to-head strength)."),
        ("MAS_%", "Expected performance into the current meta: weighted average win chance."),
        ("SE_%", "Uncertainty margin on MAS_%: higher means sparse or more variable data."),
        ("LB_%", "Conservative estimate: MAS_% - z*SE_% (z about 1.2). Penalizes sparse evidence."),
        ("BT_%", "Head-to-head strength from a Bradley-Terry model robust to missing edges."),
        ("Coverage_%", "Observed matchup coverage: percentage of total opponents faced."),
        ("N_eff", "Total considered volume: sum of W+L over all opponents."),
        ("Opp_used / Opp_total", "Distinct observed opponents / total opponents in the report."),
    ]
    ranking_legend_df = pd.DataFrame(ranking_legend_rows, columns=["Field", "Description"])
    ranking_legend_df["Color"] = pd.NA

    # Color legend.
    color_rows = [
        {"Field": "|gap_pp| >= 8",     "Description": "Large adjustment, inspect closely",  "Color": "RED"},
        {"Field": "4 <= |gap_pp| < 8", "Description": "Moderate adjustment",                "Color": "YELLOW"},
        {"Field": "Mirror",           "Description": "Self-match row",                     "Color": "GRAY"},
    ]
    if include_mas_contrib_col:
        color_rows.insert(2, {"Field": "Top-K MAS_contrib_pp", "Description": "Main contributions (K=5)", "Color": "GREEN"})
    legend_colors = pd.DataFrame(color_rows, columns=["Field", "Description", "Color"])

    legend_df = pd.concat(
        [
            legend_df,
            pd.DataFrame([{"Field": "", "Description": "Ranking legend (01_Summary)", "Color": pd.NA}]),
            ranking_legend_df,
            pd.DataFrame([{"Field": "", "Description": "Color legend", "Color": pd.NA}]),
            legend_colors,
        ],
        ignore_index=True
    )

    meta = {
        "T": len(axis),
        "K_used": float(K_used),
        "mu": float(mu),
        "gamma": (None if gamma is None else float(gamma)),
        "axis": axis,
        "global_order": order,
        "include_self_row": include_self_row,
    }
    return sheets, legend_df, meta


# ──────────────────────────────────────────────────────────────────────────────
# Summary + Workbook (data-only)
# ──────────────────────────────────────────────────────────────────────────────
def build_summary_sheet(ranking_df: pd.DataFrame) -> pd.DataFrame:
    """Build the 01_Summary sheet from the MARS ranking, ordered by Score_%."""
    keep_cols = ["Deck", "Score_%", "MAS_%", "LB_%", "BT_%", "SE_%", "N_eff", "Opp_used", "Opp_total", "Coverage_%"]
    cols = [c for c in keep_cols if c in ranking_df.columns]
    df = (ranking_df.loc[:, cols].copy()
          .sort_values("Score_%", ascending=False, kind="mergesort")
          .reset_index(drop=True))
    for c in ("Score_%", "MAS_%", "LB_%", "BT_%", "SE_%", "Coverage_%"):
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").round(2)
    return df


def prepare_workbook(
    sheets_by_deck: "OrderedDict[str, pd.DataFrame]",
    legend_df: pd.DataFrame,
    summary_df: Optional[pd.DataFrame] = None,
    *,
    include_legend_table: bool = False,
) -> "OrderedDict[str, pd.DataFrame]":
    """
    Insert '00_Legenda' as first sheet, empty when include_legend_table=False.
    If provided, insert '01_Summary' second, then one sheet per deck A.
    """
    workbook: "OrderedDict[str, pd.DataFrame]" = OrderedDict()
    workbook["00_Legenda"] = legend_df.copy() if include_legend_table else pd.DataFrame({"": []})
    if summary_df is not None:
        workbook["01_Summary"] = summary_df.copy()
    for deck, df in sheets_by_deck.items():
        workbook[_sanitize_sheet_name(deck)] = df
    return workbook


# ──────────────────────────────────────────────────────────────────────────────
# Banner (PNG) + embed on 00_Legenda
# ──────────────────────────────────────────────────────────────────────────────
def _render_legend_banner_png(
    legend_df: pd.DataFrame,
    png_path: Path,
    *,
    width: int = 1500,        # ~ larghezza dello screenshot allegato (modificabile)
    margin: int = 60
) -> Path:
    """
    Banner PNG in layout VERTICALE:
      1) What this is (callout)
      2) 01_Summary (ranking)
      3) Per-deck sheets (A -> all)
      4) Color legend

    - Deduplicate 'Mirror' in the per-deck section, keeping the first occurrence.
    - Dynamic text wrapping; soft palette.
    """
    from PIL import Image, ImageDraw, ImageFont
    import matplotlib.font_manager as fm

    # Palette
    COLS = {
        "RED": "#F2CBCB",
        "YELLOW": "#FFF2CC",
        "GREEN": "#D9EAD3",
        "GRAY": "#CDCDCD",
        "TEXT": "#222222",
        "SUB": "#444444",
        "BG": "#FFFFFF",
        "CALLOUT": "#F7F7F7",
        "BORDER": "#DDDDDD",
    }

    # Font loader
    def _load_font(weight="regular", size=18):
        fp = fm.findfont(
            fm.FontProperties(
                family="DejaVu Sans",
                weight=("bold" if weight == "bold" else "regular"),
            )
        )
        return ImageFont.truetype(fp, size=size)

    f_title = _load_font("bold", 36)
    f_h2    = _load_font("bold", 24)
    f_key   = _load_font("bold", 18)
    f_txt   = _load_font("regular", 18)
    f_meta  = _load_font("regular", 16)

    # Canvas (height grande, poi crop)
    W = int(width)
    H = 3000
    M = int(margin)
    img = Image.new("RGB", (W, H), COLS["BG"])
    d = ImageDraw.Draw(img)

    # Wrap helper
    def draw_wrapped(text, font, x, y, max_w, color, line_spacing=6):
        if not text:
            return y
        words = str(text).split()
        lines, cur = [], ""
        for w in words:
            test = (cur + " " + w).strip()
            if d.textlength(test, font=font) <= max_w:
                cur = test
            else:
                if cur:
                    lines.append(cur)
                cur = w
        if cur:
            lines.append(cur)
        ascent, descent = font.getmetrics()
        line_h = ascent + descent + line_spacing
        for ln in lines:
            d.text((x, y), ln, font=font, fill=color)
            y += line_h
        return y

    # Extract data from legend_df.
    get = lambda field: legend_df.loc[
        legend_df["Field"].astype(str) == field, "Description"
    ]
    cover_text = (get("What this is").iloc[0] if not get("What this is").empty else "")

    pr = get("Run parameters")
    param_text = (pr.iloc[0] if not pr.empty else None)

    desc = legend_df["Description"].fillna("").astype(str)
    is_sep_rank  = desc.str.startswith("Ranking legend")
    is_sep_color = desc.eq("Color legend")
    idx_rank  = is_sep_rank[is_sep_rank].index.tolist()
    idx_color = is_sep_color[is_sep_color].index.tolist()

    # 01_Summary items
    ranking_items = []
    if idx_rank and idx_color and idx_color[0] > idx_rank[0]:
        ranking_items = (
            legend_df.iloc[idx_rank[0] + 1 : idx_color[0]][["Field", "Description"]]
            .dropna(how="all")
            .to_dict("records")
        )

    # Per-deck items: only effective fields, deduplicated and ordered.
    per_deck_order = [
        "Units", "W,L,N", "WR_real_%", "p_hat_%", "SE_binom_%", "gap_pp", "Mirror", "Conventions"
    ]
    if "SE_dir_%" in legend_df["Field"].values:
        per_deck_order.insert(4, "SE_dir_%")
    if "w_A(B)_% " in legend_df["Field"].values or "w_A(B)_%" in legend_df["Field"].values:
        per_deck_order.insert(-2, "w_A(B)_%")
    if "MAS_contrib_pp" in legend_df["Field"].values:
        per_deck_order.insert(-2, "MAS_contrib_pp")

    per_deck_df = legend_df[legend_df["Field"].isin(per_deck_order)][["Field", "Description"]]
    # Deduplicate, keeping the first occurrence.
    per_deck_df = per_deck_df.drop_duplicates(subset=["Field"], keep="first")
    per_deck_df["order"] = per_deck_df["Field"].apply(lambda c: per_deck_order.index(c))
    per_deck_df = per_deck_df.sort_values("order").drop(columns="order")
    per_deck_items = per_deck_df.to_dict("records")

    # Colors.
    color_block = legend_df.loc[
        legend_df["Color"].isin(["RED", "YELLOW", "GREEN", "GRAY"]),
        ["Field", "Description", "Color"],
    ]

    # Vertical layout.
    x = M
    y = M
    max_w = W - 2 * M

    # Title.
    d.text((x, y), "Legend - how to read the per-deck report", font=f_title, fill=COLS["TEXT"])
    y += f_title.getmetrics()[0] + f_title.getmetrics()[1] + 16
    if param_text:
        y = draw_wrapped(f"Run parameters: {param_text}", f_meta, x, y, max_w, COLS["SUB"], 4) + 8

    # 1) What this is (callout).
    d.text((x, y), "What this is", font=f_h2, fill=COLS["TEXT"])
    y += f_h2.getmetrics()[0] + f_h2.getmetrics()[1] + 10
    box_h = 140
    d.rounded_rectangle([x, y, x + max_w, y + box_h], radius=14, fill=COLS["CALLOUT"], outline=COLS["BORDER"])
    draw_wrapped(cover_text, f_txt, x + 16, y + 14, max_w - 32, COLS["TEXT"])
    y = y + box_h + 18

    # 2) 01_Summary
    d.text((x, y), "01_Summary (ranking)", font=f_h2, fill=COLS["TEXT"])
    y += f_h2.getmetrics()[0] + f_h2.getmetrics()[1] + 10
    key_w = 260
    for row in ranking_items:
        field, descr = str(row["Field"]), str(row["Description"])
        d.text((x, y), field, font=f_key, fill=COLS["TEXT"])
        y = draw_wrapped(descr, f_txt, x + key_w, y, max_w - key_w, COLS["TEXT"]) + 6

    y += 6

    # 3) Per-deck sheets.
    d.text((x, y), "Per-deck sheets (A -> all)", font=f_h2, fill=COLS["TEXT"])
    y += f_h2.getmetrics()[0] + f_h2.getmetrics()[1] + 10
    for row in per_deck_items:
        field, descr = str(row["Field"]), str(row["Description"])
        d.text((x, y), field, font=f_key, fill=COLS["TEXT"])
        y = draw_wrapped(descr, f_txt, x + key_w, y, max_w - key_w, COLS["TEXT"]) + 6

    y += 6

    # 4) Color legend.
    d.text((x, y), "Color legend", font=f_h2, fill=COLS["TEXT"])
    y += f_h2.getmetrics()[0] + f_h2.getmetrics()[1] + 12
    sw_w, sw_h = 68, 36
    for _, r in color_block.iterrows():
        label, descr, key = r["Field"], r["Description"], r["Color"]
        d.rounded_rectangle([x, y, x + sw_w, y + sw_h], radius=8, fill=COLS.get(key, "#EEEEEE"), outline="#999999")
        d.text((x + sw_w + 14, y), str(label), font=f_key, fill=COLS["TEXT"])
        y = draw_wrapped(str(descr), f_txt, x + sw_w + 14, y + f_key.getmetrics()[0] + f_key.getmetrics()[1] + 6,
                         max_w - (sw_w + 14), COLS["TEXT"]) + 12

    # Final crop to used height.
    used_h = min(H, y + M)
    img = img.crop((0, 0, W, used_h))
    png_path.parent.mkdir(parents=True, exist_ok=True)
    img.save(png_path)
    return png_path


def _embed_banner_on_legend(
    excel_path: Path | str,
    png_path: Path,
    *,
    sheet_name: str = "00_Legenda",
    rows_padding: int = 36,
    wipe_sheet: bool = True,
) -> None:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage

    p = Path(excel_path)
    wb = load_workbook(p)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.worksheets[0]

    if wipe_sheet:
        for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for c in r:
                c.value = None
    if rows_padding > 0:
        ws.insert_rows(1, amount=int(rows_padding))

    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.column_dimensions["A"].width = 180
    ws.sheet_properties.tabColor = "4F81BD"
    ws.add_image(XLImage(str(png_path)), "A1")
    wb.save(p)


# ──────────────────────────────────────────────────────────────────────────────
# Riordino fogli post-scrittura (robusto)
# ──────────────────────────────────────────────────────────────────────────────
def _reorder_excel_sheets_robust(excel_path: Path | str, desired_decks: list[str]) -> None:
    """Reorder sheets by desired_decks, preserving 00_Legenda and 01_Summary first."""
    from openpyxl import load_workbook

    p = Path(excel_path)
    wb = load_workbook(p)
    existing_titles = [ws.title for ws in wb.worksheets]
    title2ws = {ws.title: ws for ws in wb.worksheets}

    def sanitize(s: str) -> str:
        return _sanitize_sheet_name(s)

    used = set()
    mapped_titles: list[str] = []

    for fixed in ("00_Legenda", "01_Summary"):
        if fixed in title2ws:
            mapped_titles.append(fixed); used.add(fixed)

    desired_san = [sanitize(d) for d in desired_decks]
    for d in desired_san:
        exact = [t for t in existing_titles if t == d and t not in used]
        if exact:
            mapped_titles.append(exact[0]); used.add(exact[0]); continue
        starts = [t for t in existing_titles if t.startswith(d) and t not in used]
        if starts:
            mapped_titles.append(starts[0]); used.add(starts[0]); continue
        lower = d.lower()
        contains = [t for t in existing_titles if (lower in t.lower()) and t not in used]
        if contains:
            mapped_titles.append(contains[0]); used.add(contains[0]); continue
        LOGGER.warning("Reorder: no sheet found for '%s' after sanitization.", d)

    for t in existing_titles:
        if t not in mapped_titles:
            mapped_titles.append(t)

    wb._sheets = [title2ws[t] for t in mapped_titles]  # openpyxl: riordino
    wb.save(p)


# ──────────────────────────────────────────────────────────────────────────────
# End-to-end: write Excel, embed banner, reorder sheets
# ──────────────────────────────────────────────────────────────────────────────
def write_mars_matchup_report(
    *,
    ranking_df: pd.DataFrame,          # must contain at least: Deck, Score_%
    filtered_wr: pd.DataFrame,
    n_dir: pd.DataFrame,
    p_blend: pd.Series,
    K_used: float,
    score_flat: Optional[pd.DataFrame] = None,
    mu: float = 0.5,
    gamma: Optional[float] = None,
    include_posterior_se: bool = False,      # default OFF
    include_binom_se: bool = True,
    include_counts: bool = True,
    include_self_row: bool = True,
    include_weight_col: bool = False,
    include_mas_contrib_col: bool = False,
    out_dir: Path | str = "outputs/reports/mars",
    base_name: Optional[str] = None,   # default: mars_matchup_report
) -> Tuple[Path, Path, Dict]:
    """
    Generate per-deck sheets, add 00_Legenda (banner only) and 01_Summary,
    write the Excel workbook (versioned + latest), embed the banner, and reorder
    sheets by ranking.
    """
    # Ordine del ranking (top→bottom)
    if not {"Deck", "Score_%"} <= set(ranking_df.columns):
        raise ValueError("ranking_df must contain at least the columns: 'Deck' and 'Score_%'.")
    ranking_order: list[str] = (
        ranking_df.sort_values("Score_%", ascending=False, kind="mergesort")["Deck"].astype(str).tolist()
    )

    # Tables and legend.
    sheets_by_deck, legend_df, meta = make_pairs_by_deck_tables(
        filtered_wr=filtered_wr,
        n_dir=n_dir,
        p_blend=p_blend,
        K_used=K_used,
        score_flat=score_flat,
        mu=mu,
        include_posterior_se=include_posterior_se,
        include_binom_se=include_binom_se,
        gamma=gamma,
        include_counts=include_counts,
        global_order=ranking_order,
        include_self_row=include_self_row,
        include_weight_col=include_weight_col,
        include_mas_contrib_col=include_mas_contrib_col,
    )

    # Summary
    summary_df = build_summary_sheet(ranking_df)

    # Workbook with empty '00_Legenda'; PNG is embedded later.
    workbook = prepare_workbook(sheets_by_deck, legend_df, summary_df, include_legend_table=False)

    # Naming & scrittura
    out_dir = Path(out_dir); out_dir.mkdir(parents=True, exist_ok=True)
    if base_name is None:
        base_name = "mars_matchup_report"

    # Writer robusto
    res = None
    try:
        res = write_excel_versioned_styled(
            workbook=workbook,
            base_dir=out_dir,
            prefix=base_name,
            tag=None,
            include_latest=True,
            also_versioned=True,
            top_k_contrib=5,
        )
    except TypeError as e:
        LOGGER.warning("write_excel_versioned_styled signature mismatch: %s - using base fallback.", e)

    if not (isinstance(res, tuple) and len(res) == 2):
        from reporting.excel import write_excel_versioned
        versioned_path, latest_path = write_excel_versioned(
            workbook=workbook,
            base_dir=out_dir,
            prefix=base_name,
            tag=None,
            include_latest=True,
            also_versioned=True,
        )
    else:
        versioned_path, latest_path = res

    # Banner on 00_Legenda.
    banner_png = _render_legend_banner_png(legend_df, out_dir / "legend_latest.png")
    for p in (versioned_path, latest_path):
        if p:
            _embed_banner_on_legend(p, banner_png, sheet_name="00_Legenda", rows_padding=36)

    # Riordino fogli secondo ranking
    _reorder_excel_sheets_robust(versioned_path, ranking_order)
    _reorder_excel_sheets_robust(latest_path, ranking_order)

    LOGGER.debug("Report scritto e riordinato | versioned=%s | latest=%s", versioned_path, latest_path)
    return Path(versioned_path), Path(latest_path), meta


def write_pairs_by_deck_report(**kwargs) -> Tuple[Path, Path, Dict]:
    """Compatibility wrapper for the old public name."""
    return write_mars_matchup_report(**kwargs)
