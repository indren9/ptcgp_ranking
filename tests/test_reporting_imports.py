import pandas as pd


def test_reporting_excel_imports_without_cycles():
    import config.loader
    import mars.report
    import reporting.excel
    import reporting.plots
    import reporting.tables
    import storage.writers
    import utils.config
    import utils.io

    assert utils.config.read_config is config.loader.read_config
    assert utils.config.get_set_params is config.loader.get_set_params
    assert mars.report.write_excel_versioned_styled is reporting.excel.write_excel_versioned_styled
    assert callable(reporting.plots.show_wr_heatmap)
    assert callable(reporting.tables.show_ranking)
    assert callable(storage.writers.write_excel_versioned_styled)
    assert utils.io.write_excel_versioned_styled is reporting.excel.write_excel_versioned_styled


def test_reporting_excel_styled_writer_smoke(tmp_path):
    from reporting.excel import write_excel_versioned_styled

    workbook = {
        "00_Legenda": pd.DataFrame(
            {
                "Campo": ["gap_pp"],
                "Descrizione": ["Differenza in punti percentuali"],
                "Colore": ["RED"],
            }
        ),
        "Deck A": pd.DataFrame(
            {
                "Opponent": ["Mirror", "Deck B"],
                "gap_pp": [0, "8.5"],
                "MAS_contrib_pp": [0, 12],
            }
        ),
    }

    versioned, latest = write_excel_versioned_styled(
        workbook,
        tmp_path,
        "pairs_by_deck",
        include_latest=True,
        also_versioned=False,
    )

    assert versioned is None
    assert latest == tmp_path / "pairs_by_deck_latest.xlsx"
    assert latest.exists()


def test_reporting_excel_styles_summary_sheet(tmp_path):
    from openpyxl import load_workbook
    from reporting.excel import write_excel_versioned_styled

    workbook = {
        "00_Legenda": pd.DataFrame({"Campo": ["gap_pp"], "Descrizione": ["Gap"], "Colore": ["RED"]}),
        "01_Summary": pd.DataFrame(
            {
                "Deck": ["Pikachu"],
                "Score_%": [55.12],
                "MAS_%": [54.0],
                "LB_%": [50.0],
                "BT_%": [58.0],
                "SE_%": [2.5],
                "N_eff": [123],
                "Opp_used": [8],
                "Opp_total": [10],
                "Coverage_%": [80.0],
            }
        ),
    }

    _, latest = write_excel_versioned_styled(
        workbook,
        tmp_path,
        "summary_report",
        include_latest=True,
        also_versioned=False,
    )

    wb = load_workbook(latest)
    ws = wb["01_Summary"]

    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:J2"
    assert ws.column_dimensions["A"].width == 34
    assert ws.column_dimensions["B"].width == 12
    assert ws["B2"].number_format == "0.00"
    assert ws["G2"].number_format == "0"
    assert ws["A1"].font.bold is True


def test_reporting_excel_styles_per_deck_sheet(tmp_path):
    from openpyxl import load_workbook
    from reporting.excel import write_excel_versioned_styled

    workbook = {
        "Deck A": pd.DataFrame(
            {
                "Opponent": ["Mirror", "Deck B"],
                "W": [None, 10],
                "L": [None, 5],
                "N": [None, 15],
                "WR_real_%": [None, 66.67],
                "p_hat_%": [None, 63.33],
                "SE_binom_%": [None, 12.17],
                "gap_pp": [None, -3.34],
            }
        )
    }

    _, latest = write_excel_versioned_styled(
        workbook,
        tmp_path,
        "deck_report",
        include_latest=True,
        also_versioned=False,
    )

    wb = load_workbook(latest)
    ws = wb["Deck A"]

    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == "A1:H3"
    assert ws.column_dimensions["A"].width == 34
    assert ws.column_dimensions["E"].width == 13
    assert ws["B3"].number_format == "0"
    assert ws["E3"].number_format == "0.00"
    assert ws["H3"].number_format == "0.00"


def test_report_legend_cover_sheet_is_presentation_friendly(tmp_path):
    from openpyxl import Workbook, load_workbook
    from PIL import Image
    from mars.report import _embed_banner_on_legend

    workbook_path = tmp_path / "report.xlsx"
    image_path = tmp_path / "legend.png"

    wb = Workbook()
    ws = wb.active
    ws.title = "00_Legenda"
    ws["A1"] = "old"
    wb.save(workbook_path)

    Image.new("RGB", (120, 60), "white").save(image_path)
    _embed_banner_on_legend(workbook_path, image_path, rows_padding=2)

    wb = load_workbook(workbook_path)
    ws = wb["00_Legenda"]

    assert ws.sheet_view.showGridLines is False
    assert ws.sheet_view.zoomScale == 85
    assert ws.column_dimensions["A"].width == 180
    assert ws.sheet_properties.tabColor.rgb == "004F81BD"


def test_report_legend_banner_accepts_field_description_schema(tmp_path):
    from mars.report import _render_legend_banner_png

    legend_df = pd.DataFrame(
        [
            {"Field": "What this is", "Description": "Deck ranking report", "Color": pd.NA},
            {"Field": "Run parameters", "Description": "K=2.0", "Color": pd.NA},
            {"Field": "", "Description": "Ranking legend (01_Summary)", "Color": pd.NA},
            {"Field": "Score_%", "Description": "Final ranking score", "Color": pd.NA},
            {"Field": "", "Description": "Color legend", "Color": pd.NA},
            {"Field": "Mirror", "Description": "Self-match row", "Color": "GRAY"},
        ]
    )

    out = _render_legend_banner_png(legend_df, tmp_path / "legend.png")

    assert out.exists()
