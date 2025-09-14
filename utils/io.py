# =============================================
# utils/io.py — routing & versioned writes (CSV + plots + Excel)
# =============================================

from __future__ import annotations

from pathlib import Path
from dataclasses import dataclass
from typing import Optional, Dict, Iterable, Tuple

import hashlib
import time
import logging
import pandas as pd
from datetime import datetime

# Minimal logger (fallback); in runtime potresti avere già "ptcgp"
log = logging.getLogger("ptcgp") if logging.getLogger("ptcgp").handlers else logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────────────────────
# Paths & init
# ──────────────────────────────────────────────────────────────────────────────

@dataclass
class Paths:
    base: Path
    outputs: Path
    cache: Path
    logs: Path


def init_paths(base_dir: Path) -> Paths:
    """
    Crea (idempotentemente) la struttura minima di cartelle usata dal progetto.
    """
    base = Path(base_dir)
    out = base / "outputs"
    cache = base / "cache" / "requests"
    logs = base / "logs"

    for d in [
        out / "Decklists" / "raw",
        out / "Decklists" / "top_meta",
        out / "MatchupData" / "raw",
        out / "MatchupData" / "flat",
        out / "Matrices" / "winrate",
        out / "Matrices" / "volumes",
        out / "Matrices" / "heatmap",
        out / "RankingData" / "MARS",
        out / "RankingData" / "MARS" / "archives",
        out / "RankingData" / "MARS" / "Report",   # cartella report Excel
        cache,
        logs,
    ]:
        d.mkdir(parents=True, exist_ok=True)

    return Paths(base=base, outputs=out, cache=cache, logs=logs)


# ──────────────────────────────────────────────────────────────────────────────
# Routing
# ──────────────────────────────────────────────────────────────────────────────

ROUTES: dict[str, tuple[str, ...]] = {
    # CSV contract
    "decklist_raw": ("Decklists", "raw"),
    "top_meta_decklist": ("Decklists", "top_meta"),
    "matchup_score_table": ("MatchupData", "flat"),
    "filtered_wr": ("Matrices", "winrate"),
    "n_dir": ("Matrices", "volumes"),

    # Top-meta già post-alias
    "top_meta_post_alias": ("Decklists", "top_meta"),

    # Plots
    "heatmap_topN": ("Matrices", "heatmap"),

    # MARS outputs
    "mars_ranking": ("RankingData", "MARS"),

    # Report per-deck (Excel multi-sheet)
    "report": ("RankingData", "MARS", "Report"),
}


def _dest(paths: Paths, prefix: str) -> Path:
    """
    Risolve il percorso di destinazione per un dato 'prefix' di ROUTES.
    Supporta route con 2 o più segmenti.
    """
    if prefix not in ROUTES:
        log.warning("[route] Prefix sconosciuto '%s' — invio a outputs/", prefix)
        return paths.outputs
    parts = ROUTES[prefix]
    dest = paths.outputs
    for p in parts:
        dest = dest / p
    return dest


def _run_stamp() -> str:
    """Timestamp run-wide in formato YYYYmmdd_HHMMSS."""
    return time.strftime("%Y%m%d_%H%M%S")


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _df_content_hash(df: pd.DataFrame) -> str:
    """
    Hash stabile del contenuto di un DataFrame (CSV bytes senza index).
    Utile per decidere se scrivere una copia versionata.
    """
    as_csv = df.to_csv(index=False).encode("utf-8")
    import hashlib as _hashlib  # locale per evitare shadowing
    return _hashlib.sha256(as_csv).hexdigest()[:16]


# ──────────────────────────────────────────────────────────────────────────────
# CSV writer
# ──────────────────────────────────────────────────────────────────────────────

def write_csv_versioned(
    df: pd.DataFrame,
    base_dir: Path | str,
    prefix: str,
    *,
    changed: bool,
    index: bool = False,
) -> Path:
    """
    Scrive sempre <prefix>_latest.csv e, se changed=True, aggiunge anche
    una copia versionata <prefix>_<timestamp>.csv.
    Ritorna il Path del file scritto più recentemente (versionato se created, altrimenti latest).
    """
    dest_dir = Path(base_dir)
    dest_dir.mkdir(parents=True, exist_ok=True)

    latest_path = dest_dir / f"{prefix}_latest.csv"
    df.to_csv(latest_path, index=index, encoding="utf-8")
    log.info("CSV aggiornato: %s", latest_path)

    if changed:
        ts = _run_stamp()
        ts_path = dest_dir / f"{prefix}_{ts}.csv"
        df.to_csv(ts_path, index=index, encoding="utf-8")
        log.info("CSV versionato (changed=True): %s", ts_path)
        return ts_path

    return latest_path


# ──────────────────────────────────────────────────────────────────────────────
# Plot writers
# ──────────────────────────────────────────────────────────────────────────────

def save_plot_timestamped(fig, base_dir: Path | str, prefix: str, *, fmt: str = "png", dpi: int = 300) -> Path:
    """
    Salva SEMPRE un'immagine con timestamp: <prefix>_<timestamp>.<fmt>
    (Manteniamo per retro-compatibilità.)
    """
    base_dir = Path(base_dir)
    base_dir.mkdir(parents=True, exist_ok=True)
    ts = _run_stamp()
    path = base_dir / f"{prefix}_{ts}.{fmt}"
    fig.savefig(path, dpi=dpi, bbox_inches="tight")
    log.debug("Plot salvato (timestamped): %s", path)
    return path


def save_plot_dual(fig, base_dir: Path | str, prefix: str, tag: str, *, fmt: str = "png", dpi: int = 300) -> tuple[Path, Path]:
    """
    Salva DUE copie del plot:
      1) <prefix>_latest.<fmt>              (sovrascritto ad ogni run)
      2) <prefix>_<tag>_<timestamp>.<fmt>   (versionato con timestamp)

    Esempio:
      save_plot_dual(fig, out_dir, "wr_heatmap", tag="T15")
      → wr_heatmap_latest.png
        wr_heatmap_T15_20250101_120000.png

    Ritorna (ts_path, latest_path).
    """
    base_dir = Path(base_dir)
    base_dir.mkdir(parents=True, exist_ok=True)

    ts = _run_stamp()
    ts_path = base_dir / f"{prefix}_{tag}_{ts}.{fmt}"
    latest_path = base_dir / f"{prefix}_latest.{fmt}"

    fig.savefig(ts_path, dpi=dpi, bbox_inches="tight")
    fig.savefig(latest_path, dpi=dpi, bbox_inches="tight")
    log.info("Plot salvato (timestamp + latest): %s | latest: %s", ts_path, latest_path)
    return ts_path, latest_path


# ──────────────────────────────────────────────────────────────────────────────
# Excel writer (base)
# ──────────────────────────────────────────────────────────────────────────────

def write_excel_versioned(
    workbook: "dict[str, pd.DataFrame]",
    base_dir: Path | str,
    prefix: str,
    *,
    tag: str | None = None,
    include_latest: bool = True,
    also_versioned: bool = True,
) -> tuple[Path | None, Path | None]:
    """
    Scrive un Excel multi-sheet mantenendo l'ordine dei fogli fornito.
    Ritorna (ts_path, latest_path). Se una delle due copie non viene scritta, ritorna None su quella.
    """
    dest_dir = Path(base_dir)
    dest_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    tag_part = f"_{tag}" if tag else ""
    ts_name = f"{prefix}{tag_part}_{ts}.xlsx"
    latest_name = f"{prefix}_latest.xlsx"

    ts_path = dest_dir / ts_name if also_versioned else None
    latest_path = dest_dir / latest_name if include_latest else None

    # Seleziona engine disponibile
    engine = None
    for eng in ("openpyxl", "xlsxwriter"):
        try:
            __import__(eng)
            engine = eng
            break
        except Exception:
            continue
    if engine is None:
        engine = "openpyxl"

    def _write(path: Path) -> None:
        with pd.ExcelWriter(path, engine=engine) as xw:
            for sheet_name, df in workbook.items():
                df.to_excel(xw, sheet_name=sheet_name, index=False)

    if ts_path is not None:
        _write(ts_path)
        log.info("Excel versionato: %s", ts_path)
    if latest_path is not None:
        _write(latest_path)
        log.info("Excel latest: %s", latest_path)

    return ts_path, latest_path


# ──────────────────────────────────────────────────────────────────────────────
# Excel writer con styling (semafori gap, Top-K, Mirror, legenda colori)
# ──────────────────────────────────────────────────────────────────────────────

def write_excel_versioned_styled(
    workbook: "dict[str, pd.DataFrame]",
    base_dir: "Path | str",
    prefix: str,
    *,
    tag: "str | None" = None,
    include_latest: bool = True,
    also_versioned: bool = True,
    top_k_contrib: int = 5,
) -> "tuple[Path | None, Path | None]":
    """
    Come write_excel_versioned, ma aggiunge:
      - semafori su gap_pp (rosso |gap|>=8; giallo 4–8) via FormulaRule
      - evidenziazione Top-K su MAS_contrib_pp con Rule(type="top10") nativa
      - colorazione riga 'Mirror' (grigio #CDCDCD) + 'Opponent' in corsivo
      - swatch di colore nella colonna 'Colore' del foglio 00_Legenda
    Compatibile con openpyxl 3.1.x (usa Rule + DifferentialStyle).
    """
    from datetime import datetime
    from pathlib import Path
    import pandas as pd

    dest_dir = Path(base_dir)
    dest_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    tag_part = f"_{tag}" if tag else ""
    ts_name = f"{prefix}{tag_part}_{ts}.xlsx"
    latest_name = f"{prefix}_latest.xlsx"

    ts_path = dest_dir / ts_name if also_versioned else None
    latest_path = dest_dir / latest_name if include_latest else None

    engine = "openpyxl"

    def _write_plain(path: Path) -> None:
        with pd.ExcelWriter(path, engine=engine) as xw:
            for sheet_name, df in workbook.items():
                df.to_excel(xw, sheet_name=sheet_name, index=False)

    def _style_in_place(path: Path) -> None:
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import PatternFill, Font
            from openpyxl.formatting.rule import FormulaRule, Rule
            from openpyxl.styles.differential import DifferentialStyle
        except Exception as e:
            log.warning("openpyxl non disponibile per styling: %s", e)
            return

        wb = load_workbook(path)

        for ws in wb.worksheets:
            name = ws.title

            # Header map
            headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1) if isinstance(cell.value, str)}
            max_row = ws.max_row
            max_col = ws.max_column
            if max_row < 2:
                continue

            # ===== Styling specifico per 00_Legenda: colonna 'Colore' come swatch =====
            if name == "00_Legenda" and "Colore" in headers:
                col_c = headers["Colore"]

                # mapping chiave -> colore (ARGB)
                cmap = {
                    "RED":   "FFF2CBCB",  # rosso chiaro
                    "YELLOW":"FFFFF2CC",  # giallo chiaro
                    "GREEN": "FFD9EAD3",  # verde chiaro
                    "GRAY":  "FFCDCDCD",  # grigio #CDCDCD
                }

                for r in range(2, max_row + 1):
                    key_cell = ws.cell(row=r, column=col_c)
                    key = key_cell.value
                    if isinstance(key, str):
                        key_norm = key.strip().upper()
                        if key_norm in cmap:
                            key_cell.fill = PatternFill(start_color=cmap[key_norm], end_color=cmap[key_norm], fill_type="solid")
                            # opzionale: rimuovi il testo e lascia solo lo swatch
                            key_cell.value = ""
                # Non applicare altro styling alla legenda
                continue

            # ===== Fogli per-deck / Summary =====

            # Semafori su gap_pp: rosso |gap|>=8; giallo 4–8
            if "gap_pp" in headers:
                col = headers["gap_pp"]
                L = get_column_letter(col)
                rng = f"{L}2:{L}{max_row}"

                red_fill = PatternFill(start_color="FFF2CBCB", end_color="FFF2CBCB", fill_type="solid")
                yellow_fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

                ws.conditional_formatting.add(rng, FormulaRule(formula=[f"ABS({L}2)>=8"], fill=red_fill))
                ws.conditional_formatting.add(rng, FormulaRule(formula=[f"AND(ABS({L}2)>=4,ABS({L}2)<8)"], fill=yellow_fill))

            # Top-K nativo su MAS_contrib_pp
            if "MAS_contrib_pp" in headers and top_k_contrib and name not in ("00_Legenda", "01_Summary"):
                col = headers["MAS_contrib_pp"]
                L = get_column_letter(col)
                rng = f"{L}2:{L}{max_row}"

                dxf = DifferentialStyle(fill=PatternFill(start_color="FFD9EAD3",
                                                         end_color="FFD9EAD3",
                                                         fill_type="solid"))
                rule = Rule(type="top10", rank=int(top_k_contrib), percent=False, bottom=False, dxf=dxf)
                ws.conditional_formatting.add(rng, rule)

            # Riga 'Mirror' in grigio su tutta la riga + Opponent in corsivo
            opp_col = headers.get("Opponent")
            if opp_col is not None and name not in ("00_Legenda", "01_Summary"):
                gray = PatternFill(start_color="FFCDCDCD", end_color="FFCDCDCD", fill_type="solid")
                for r in range(2, max_row + 1):
                    cell = ws.cell(row=r, column=opp_col)
                    if isinstance(cell.value, str) and cell.value.strip().lower() == "mirror":
                        # Italic solo per la cella Opponent
                        try:
                            cell.font = Font(name=cell.font.name, sz=cell.font.sz, bold=cell.font.bold,
                                             italic=True, vertAlign=cell.font.vertAlign,
                                             underline=cell.font.underline, color=cell.font.color)
                        except Exception:
                            # fallback compatibile
                            cell.font = cell.font.copy(italic=True)
                        # Grigio su tutta la riga
                        for c in range(1, max_col + 1):
                            ws.cell(row=r, column=c).fill = gray

        wb.save(path)

    # Scrivi i file richiesti e applica styling
    if ts_path is not None:
        _write_plain(ts_path)
        _style_in_place(ts_path)
        log.info("Excel versionato (styled): %s", ts_path)
    if latest_path is not None:
        _write_plain(latest_path)
        _style_in_place(latest_path)
        log.info("Excel latest (styled): %s", latest_path)

    return ts_path, latest_path

# --- utils/io.py ---

from pathlib import Path

