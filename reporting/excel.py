from __future__ import annotations

from datetime import datetime
from pathlib import Path
import logging
import time

import pandas as pd

from storage.writers import write_excel_versioned

log = logging.getLogger("ptcgp")


def write_excel_versioned_styled(
    workbook: dict[str, pd.DataFrame],
    base_dir: Path | str,
    prefix: str,
    *,
    tag: str | None = None,
    include_latest: bool = True,
    also_versioned: bool = True,
    top_k_contrib: int = 5,
) -> tuple[Path | None, Path | None]:
    """
    Write a multi-sheet Excel file with report styling.

    Adds conditional formatting for gap_pp, top-K contribution highlights,
    Mirror row styling, legend swatches, and atomic Windows-friendly writes.
    """
    import os
    import tempfile

    dest_dir = Path(base_dir)
    dest_dir.mkdir(parents=True, exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    tag_part = f"_{tag}" if tag else ""
    ts_path = dest_dir / f"{prefix}{tag_part}_{ts}.xlsx" if also_versioned else None
    latest_path = dest_dir / f"{prefix}_latest.xlsx" if include_latest else None

    engine = None
    for candidate in ("xlsxwriter", "openpyxl"):
        try:
            __import__(candidate)
            engine = candidate
            break
        except Exception:
            continue
    if engine is None:
        engine = "openpyxl"

    def _atomic_write(path: Path, *, retries: int = 6, backoff_s: float = 0.7) -> Path:
        fd, tmp_name = tempfile.mkstemp(dir=dest_dir, prefix=path.stem + "_", suffix=".tmp.xlsx")
        os.close(fd)
        tmp = Path(tmp_name)

        try:
            with pd.ExcelWriter(tmp, engine=engine) as xw:
                for sheet_name, df in workbook.items():
                    df.to_excel(xw, sheet_name=sheet_name, index=False)
        except Exception:
            try:
                tmp.unlink(missing_ok=True)
            finally:
                raise

        last_exc: Exception | None = None
        for i in range(max(1, retries)):
            try:
                os.replace(tmp, path)
                return path
            except PermissionError as e:
                last_exc = e
                time.sleep(backoff_s * (1.8 ** i))

        fallback = path.with_name(f"{path.stem}_LOCKED_{datetime.now().strftime('%Y%m%d_%H%M%S')}{path.suffix}")
        try:
            os.replace(tmp, fallback)
        except Exception:
            import shutil

            shutil.copy2(tmp, fallback)
            tmp.unlink(missing_ok=True)
        log.warning("write_excel_versioned_styled: target lockato: %s -> salvato come fallback: %s", path, fallback)
        if last_exc:
            log.debug("Ultimo errore lock Excel: %s", last_exc)
        return fallback

    def _style_in_place(path: Path, *, retries: int = 6, backoff_s: float = 0.7) -> None:
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.formatting.rule import CellIsRule, Rule
            from openpyxl.styles.differential import DifferentialStyle
        except Exception as e:
            log.warning("openpyxl is not available for styling: %s", e)
            return

        last_exc: Exception | None = None
        for i in range(max(1, retries)):
            try:
                wb = load_workbook(path)
                for ws in wb.worksheets:
                    name = ws.title
                    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1) if isinstance(cell.value, str)}
                    max_row = ws.max_row
                    max_col = ws.max_column
                    if max_row < 2:
                        continue

                    header_fill = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
                    header_font = Font(color="FFFFFFFF", bold=True)
                    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    thin_bottom = Border(bottom=Side(style="thin", color="FFB8C0CC"))

                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                        cell.border = thin_bottom
                    ws.freeze_panes = "A2"
                    ws.auto_filter.ref = ws.dimensions

                    if name == "00_Legenda":
                        field_col = "Field" if "Field" in headers else "Campo"
                        desc_col = "Description" if "Description" in headers else "Descrizione"
                        color_col = "Color" if "Color" in headers else "Colore"

                        if field_col in headers:
                            ws.column_dimensions[get_column_letter(headers[field_col])].width = 26
                        if desc_col in headers:
                            ws.column_dimensions[get_column_letter(headers[desc_col])].width = 92
                        if color_col in headers:
                            ws.column_dimensions[get_column_letter(headers[color_col])].width = 12

                        wrap_top = Alignment(wrap_text=True, vertical="top", horizontal="left")
                        for col_name in (field_col, desc_col):
                            col_idx = headers.get(col_name)
                            if col_idx is None:
                                continue
                            for r in range(1, max_row + 1):
                                ws.cell(row=r, column=col_idx).alignment = wrap_top

                        if color_col in headers:
                            col_c = headers[color_col]
                            cmap = {
                                "RED": "FFF2CBCB",
                                "YELLOW": "FFFFF2CC",
                                "GREEN": "FFD9EAD3",
                                "GRAY": "FFCDCDCD",
                            }
                            for r in range(2, max_row + 1):
                                key_cell = ws.cell(row=r, column=col_c)
                                key = key_cell.value
                                if isinstance(key, str):
                                    key_norm = key.strip().upper()
                                    if key_norm in cmap:
                                        key_cell.fill = PatternFill(
                                            start_color=cmap[key_norm],
                                            end_color=cmap[key_norm],
                                            fill_type="solid",
                                        )
                                        key_cell.value = ""
                        continue

                    if name == "01_Summary":
                        widths = {
                            "Deck": 34,
                            "Score_%": 12,
                            "MAS_%": 12,
                            "LB_%": 12,
                            "BT_%": 12,
                            "SE_%": 12,
                            "N_eff": 12,
                            "Opp_used": 12,
                            "Opp_total": 12,
                            "Coverage_%": 14,
                        }
                        percent_cols = {"Score_%", "MAS_%", "LB_%", "BT_%", "SE_%", "Coverage_%"}
                        integer_cols = {"N_eff", "Opp_used", "Opp_total"}

                        for col_name, width in widths.items():
                            col_idx = headers.get(col_name)
                            if col_idx is not None:
                                ws.column_dimensions[get_column_letter(col_idx)].width = width

                        for row in range(2, max_row + 1):
                            for col_name in percent_cols:
                                col_idx = headers.get(col_name)
                                if col_idx is not None:
                                    ws.cell(row=row, column=col_idx).number_format = "0.00"
                            for col_name in integer_cols:
                                col_idx = headers.get(col_name)
                                if col_idx is not None:
                                    ws.cell(row=row, column=col_idx).number_format = "0"

                        deck_col = headers.get("Deck")
                        if deck_col is not None:
                            for row in range(2, max_row + 1):
                                ws.cell(row=row, column=deck_col).alignment = Alignment(horizontal="left")
                        continue

                    per_deck_widths = {
                        "Opponent": 34,
                        "W": 9,
                        "L": 9,
                        "N": 9,
                        "WR_real_%": 13,
                        "p_hat_%": 13,
                        "SE_dir_%": 12,
                        "SE_binom_%": 13,
                        "gap_pp": 12,
                        "w_A(B)_%": 12,
                        "MAS_contrib_pp": 16,
                    }
                    per_deck_decimal_cols = {
                        "WR_real_%",
                        "p_hat_%",
                        "SE_dir_%",
                        "SE_binom_%",
                        "gap_pp",
                        "w_A(B)%",
                        "w_A(B)_%",
                        "w_A(B)_%", 
                        "MAS_contrib_pp",
                    }
                    per_deck_integer_cols = {"W", "L", "N"}

                    for col_name, width in per_deck_widths.items():
                        col_idx = headers.get(col_name)
                        if col_idx is not None:
                            ws.column_dimensions[get_column_letter(col_idx)].width = width

                    opponent_col = headers.get("Opponent")
                    if opponent_col is not None:
                        for row in range(2, max_row + 1):
                            ws.cell(row=row, column=opponent_col).alignment = Alignment(horizontal="left")

                    for row in range(2, max_row + 1):
                        for col_name in per_deck_decimal_cols:
                            col_idx = headers.get(col_name)
                            if col_idx is not None:
                                ws.cell(row=row, column=col_idx).number_format = "0.00"
                        for col_name in per_deck_integer_cols:
                            col_idx = headers.get(col_name)
                            if col_idx is not None:
                                ws.cell(row=row, column=col_idx).number_format = "0"

                    if "gap_pp" in headers:
                        col = headers["gap_pp"]
                        letter = get_column_letter(col)
                        cell_range = f"{letter}2:{letter}{max_row}"

                        red_fill = PatternFill(start_color="FFF2CBCB", end_color="FFF2CBCB", fill_type="solid")
                        yellow_fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

                        for row in range(2, max_row + 1):
                            cell = ws.cell(row=row, column=col)
                            if isinstance(cell.value, str):
                                try:
                                    cell.value = float(cell.value.replace(",", "."))
                                except Exception:
                                    pass

                        ws.conditional_formatting.add(
                            cell_range,
                            CellIsRule(operator="greaterThanOrEqual", formula=["8"], fill=red_fill),
                        )
                        ws.conditional_formatting.add(
                            cell_range,
                            CellIsRule(operator="lessThanOrEqual", formula=["-8"], fill=red_fill),
                        )
                        ws.conditional_formatting.add(
                            cell_range,
                            CellIsRule(operator="between", formula=["4", "8"], fill=yellow_fill),
                        )
                        ws.conditional_formatting.add(
                            cell_range,
                            CellIsRule(operator="between", formula=["-8", "-4"], fill=yellow_fill),
                        )

                    if "MAS_contrib_pp" in headers and top_k_contrib and name not in ("00_Legenda", "01_Summary"):
                        col = headers["MAS_contrib_pp"]
                        letter = get_column_letter(col)
                        cell_range = f"{letter}2:{letter}{max_row}"
                        dxf = DifferentialStyle(
                            fill=PatternFill(start_color="FFD9EAD3", end_color="FFD9EAD3", fill_type="solid")
                        )
                        rule = Rule(type="top10", rank=int(top_k_contrib), percent=False, bottom=False, dxf=dxf)
                        ws.conditional_formatting.add(cell_range, rule)

                    opp_col = headers.get("Opponent")
                    if opp_col is not None and name not in ("00_Legenda", "01_Summary"):
                        gray = PatternFill(start_color="FFCDCDCD", end_color="FFCDCDCD", fill_type="solid")
                        for row in range(2, max_row + 1):
                            cell = ws.cell(row=row, column=opp_col)
                            if isinstance(cell.value, str) and cell.value.strip().lower() == "mirror":
                                try:
                                    cell.font = Font(
                                        name=cell.font.name,
                                        sz=cell.font.sz,
                                        bold=cell.font.bold,
                                        italic=True,
                                        vertAlign=cell.font.vertAlign,
                                        underline=cell.font.underline,
                                        color=cell.font.color,
                                    )
                                except Exception:
                                    cell.font = cell.font.copy(italic=True)
                                for col in range(1, max_col + 1):
                                    ws.cell(row=row, column=col).fill = gray

                wb.save(path)
                return
            except PermissionError as e:
                last_exc = e
                time.sleep(backoff_s * (1.8 ** i))
        log.warning("Styling skipped because the file remained locked: %s (last error: %s)", path, last_exc)

    if ts_path is not None:
        ts_path = _atomic_write(ts_path)
        _style_in_place(ts_path)
        log.debug("Excel versionato (styled): %s", ts_path)
    if latest_path is not None:
        latest_path = _atomic_write(latest_path)
        _style_in_place(latest_path)
        log.debug("Excel latest (styled): %s", latest_path)

    return ts_path, latest_path


__all__ = ["write_excel_versioned", "write_excel_versioned_styled"]
